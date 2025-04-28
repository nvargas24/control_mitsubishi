from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import calendar
import locale

import pandas as pd
from dotenv import load_dotenv
import os
import re

import questionary

# Configurar el idioma a español
locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252') 

load_dotenv('urls_sarmiento.env')
url_file_mit = os.getenv("URL_MITSUBISHI")
url_desktop = os.getenv("URL_DESKTOP")

list_components_bch = ['IGB1 1', 'IGB1 2', 'IGB1', 'IGB2', 'DB1', 'DB2']
list_components_pwu = ['IGD5 U', 'IGD5 V', 'IGD5 W', 'IGU', 'IGV', 'IGW', 'IGX', 'IGY', 'IGZ']

Coches_formacion =['M1-1', 'M2-1', 'M1-2', 'M2-2', 'M3', 'M4']

def extract_xlsx():
    # Extraccion de URL del archivo xlsx desde un .env
    URL = os.path.join(url_file_mit, "FALLAS VVVF MITSUBISHI 20240129_V1.0.xlsx")

    # Carga de archivo xlsx
    wb = load_workbook(URL)
    ws = wb.active

    try:
        df_test = pd.read_excel(URL, sheet_name=ws.title, header=None)
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")

    start_table = "A1"
    end_table = "X1"

    start_row, start_col = coordinate_to_tuple(start_table)
    _, end_col = coordinate_to_tuple(end_table)

    # Ubico tabla en la hoja activa y obtengo un objeto generator
    rows = ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row, 
        min_col=start_col, 
        max_col=end_col,
    )

    # Itero sobre generator para cargar valores en una lista de listas
    data = []

    for row in rows:
        data.append([cell.value for cell in row])

    # Descrimino entre encabezado y datos
    headers = data[0]
    rows_data = data[1:]

    # Creo dataframe
    df = pd.DataFrame(rows_data, columns=headers)    

    # Identifico y filtro solo filas validas - en base a num de registro
    df["Num"] = pd.to_numeric(df["N"], errors="coerce")
    df_clean = df.dropna(subset=["Num"])
    df_clean = df_clean.drop("Num", axis=1)

    df_clean = df_clean.reset_index(drop=True)

    return df_clean

def calculate_time_between_failures(df):
    # Calcular la diferencia en días entre fechas consecutivas
    df['Tiempo entre fallas (días)'] = df['Fecha de falla'].diff().dt.days
    # Reemplazar NaN (primer registro) con 0
    df['Tiempo entre fallas (días)'] = df['Tiempo entre fallas (días)'].fillna(0).astype(int)

    return df

def search_serie(df, num_serie):
    df_filter = df.copy()

    try:
        df_filter = df_filter[df_filter["Número de serie"] == num_serie]

        # Identifico modulo segun numero de serie solicitado
        if not df_filter.empty:
            modulo = df_filter.iloc[0]['Unidad en falla']
        else:
            raise ValueError(f"No se encontró el número de serie: {num_serie}")

        # Borra columnas no relevantes
        df_filter = df_filter.drop(['Número de serie', 'Unidad en falla','UBICACIÓN ACTUAL', 'GPS', 'componentes_reemplazados'], axis=1)

        # Calculo de lapso de tiempo entre ingresos
        df_filter = calculate_time_between_failures(df_filter)

        return df_filter, modulo

    except ValueError as ve:
        print(ve)
        return pd.DataFrame(), None  # Retorna un DataFrame vacío y `None` como módulo
    except KeyError as ke:
        print(f"Error: Falta una columna esperada en el DataFrame: {ke}")
        return pd.DataFrame(), None
    except Exception as e:
        print(f"Error inesperado: {e}")
        return pd.DataFrame(), None

def filter_by_type(df, type):
    df_filter = df.copy()

    df_filter = df_filter[df_filter["Unidad en falla"] == type]
    #df_filter = df_filter.drop("Unidad en falla", axis=1)

    if type=='BCH':
        df_filter = df_filter.drop(list_components_pwu, axis=1)
    elif type=='PWU':
        df_filter = df_filter.drop(list_components_bch, axis=1)

    return df_filter

def segment_components(register):
    valores_validos = ['x', 'xp', 'p']
    list_components = []
    components = ""

    if register["Unidad en falla"].strip() == "BCH":
        list_components = list_components_bch
    elif register["Unidad en falla"].strip() == "PWU":
        list_components = list_components_pwu 

    for col in df.columns:
        row = str(register[col]).strip().lower()

        if row in valores_validos and col in list_components:
           components = components + ", "+ f"{col}" 

    components = components.strip(", ")

    if components:
        return components
    else:
        return "Sin registro"

def used_components(register):
    # Inicializar contadores
    counts = {
        "IGBT RM600HS-34S": 0,
        "Diodo CM2400HCB34N": 0,
        "Resistor 330k": 0,
        "Diodo zener 1N4746": 0,
        "Placa de control": 0
    }

    # Mapear columnas a sus categorías
    component_mapping = {
        'IGB1 1': "Placa de control",
        'IGB1 2': "Placa de control",
        'IGB1': "IGBT RM600HS-34S",
        'IGB2': "IGBT RM600HS-34S",
        'DB1': "Diodo CM2400HCB34N",
        'DB2': "Diodo CM2400HCB34N",
        'IGD5 U': "Placa de control",
        'IGD5 V': "Placa de control",
        'IGD5 W': "Placa de control",
        'IGU': "IGBT RM600HS-34S",
        'IGV': "IGBT RM600HS-34S",
        'IGW': "IGBT RM600HS-34S",
        'IGX': "IGBT RM600HS-34S",
        'IGY': "IGBT RM600HS-34S",
        'IGZ': "IGBT RM600HS-34S"
    }

    # Determinar lista de componentes según la unidad en falla
    list_components = []

    if register["Unidad en falla"].strip() == "BCH":
        list_components = list_components_bch

    elif register["Unidad en falla"].strip() == "PWU":
        list_components = list_components_pwu

    # Contar componentes válidos
    valores_validos = ['x', 'xp']
    
    for col in list_components:
        row = str(register[col]).strip().lower()
        if col in component_mapping:
            # Suma de componentes segun encabezado
            if row in valores_validos:
                category = component_mapping[col]
                counts[category] += 1
            # Suma de componentes no especificado por encabezado
            if row=='xp' or row=='p':
                counts["Diodo zener 1N4746"] += 2
                counts["Resistor 330k"] += 1

    return counts

def filter_re_code(text):
    """
    Filtra textos que contengan la estructura 'RE****'.
    Si no se encuentra, devuelve 'Sin registro'.
    """
    match = re.search(r'\bRE\d{4}\b', text)
    if match:
        return match.group(0)
    else:
        return "Sin registro"

def enrich_dataframe(df):
    # Componentes utilizados
    counts_df = df.apply(used_components, axis=1, result_type="expand")
    df = pd.concat([df, counts_df], axis=1)
    # Serigrafia de componentes en equipo
    df['componentes_reemplazados'] = df.apply(segment_components, axis=1)
    # Filtrado de Num RE de equipos
    df['Cod_Rep'] = df['UBICACIÓN ACTUAL'].apply(filter_re_code)

    df['Año'] = pd.to_datetime(df['Fecha de falla']).dt.year
    df['Mes'] = pd.to_datetime(df['Fecha de falla']).dt.strftime('%b') 

    return df

def view_serie(df, cant_reg=0):
    df_serie = df.copy()

    list_cant_sup_reg = []
    num_reg_max = 0

    for n_serie in df_serie['Número de serie'].unique():
        try:
            df_filtered, modulo = search_serie(df_serie, f"{n_serie}")
            # ##### FILTRAR SI ES PWU O BCH  ***** en todas las lecturas

            df_filtered = df_filtered.drop(['Año', 'Mes'], axis=1)
            # Procesar solo si no hay errores en search_serie
            num_reg = df_filtered.shape[0]
            
            if num_reg > num_reg_max:
                num_reg_max = num_reg

            if num_reg == cant_reg:
                print(f"\n ********************** Equipo {modulo} - {n_serie} ****************************")
                print(df_filtered)

            if num_reg > cant_reg:
                list_cant_sup_reg.append(n_serie)

        except ValueError as ve:
            print(ve)
        except Exception as e:
            print(f"Error inesperado: {e}")

    if cant_reg > num_reg_max:
        raise ValueError(f"El valor de cant_reg ({cant_reg}) es mayor que el número máximo de registros ({num_reg}) para el número de serie: {n_serie}")

    if list_cant_sup_reg:
        str_sup_reg = ", ".join(list_cant_sup_reg)

        print(f" *****  ADV: Hay modulos con un numero mayor de {cant_reg} registro/s:")
        print(f" *****  {str_sup_reg}")

def view_modulo(df, n_serie):
    df_serie = df.copy()

    df_filtered, modulo = search_serie(df_serie, f"{n_serie}")
    if not df_filtered.empty and modulo is not None:
        print(f"\n ********************** Equipo {modulo} - {n_serie} ****************************")
        print(df_filtered)

def view_component_used(df):
    df_comp = df.copy()
    
    df_resume = None  # Inicializar el DataFrame resumen
    #list_modulos = df_comp["Unidad en falla"].unique()
    list_anios = df_comp['Año'].unique()
    list_modulo = df_comp['Unidad en falla'].unique()

    for modulo in list_modulo:
        for year in list_anios:
            try:
                # Generar el DataFrame para la combinación actual
                df_result = components_used_by_month(df_comp, modulo, year)
                # Agregar la columna 'Total' con la suma de todas las columnas de meses
                df_result[year] = df_result.drop(columns=['Componente']).sum(axis=1)
                # Mantener solo las columnas 'Componente' y el total del año
                df_result = df_result[['Componente', year]]
                
                # Combinar con el DataFrame resumen
                if df_resume is None:
                    df_resume = df_result  # Inicializar con el primer DataFrame
                else:
                    # Combinar basándose en la columna 'Componente'
                    df_resume = pd.merge(df_resume, df_result, on='Componente', how='outer')

            except Exception as e:
                print(f"Error al procesar módulo '{modulo}' y año '{year}': {e}")

        df_resume = df_resume.apply(
            lambda col: col.fillna(0).astype(int) if col.dtype in ['float64', 'int64'] else col
        )
        print(f"\n ********************** Equipos {modulo} ****************************")
        print(df_resume)  
        df_resume = None


def components_used_by_month(df, unidad_falla, year):
    # Listado de meses
    orden_meses = [calendar.month_abbr[i].lower() for i in range(1, 13)] 

    df_comp = df.copy()

    df_filtered = df_comp[(df_comp['Año'] == year) & (df_comp['Unidad en falla'] == unidad_falla)]

    list_drop = ['N','Fecha de falla', 'Formación', 'Coche', 'Número de serie', 'ESTADO ACTUAL', 'UBICACIÓN ACTUAL', 'Unidad en falla', 'GPS', 'componentes_reemplazados', 'Cod_Rep', 'Año']
    df_filtered = df_filtered.drop(list_drop, axis=1)

    #if df_filtered.empty:
    #    return pd.DataFrame(columns=['Componente'] + orden_meses)


    # Listado de componentes para este modulo
    componentes = [col for col in df_filtered.columns if col != "Mes"]
    # Se crea df con columna de componetes y sus cantidades
    df_melt = df_filtered.melt(id_vars="Mes", value_vars=componentes, var_name="Componente", value_name="Cantidad")
    # Agrupa por Componente y Mes, debido a que aparecen varias veces por los distintos num_serie
    df_grouped = df_melt.groupby(['Componente', 'Mes'])['Cantidad'].sum()
    # Crea columnas de cada mes
    df_resultado = df_grouped.unstack(fill_value=0)
    # Asegurarte de que el índice no tenga nombre
    df_resultado.index.name = None

    # Reordenar las columnas de los meses
    df_resultado = df_resultado.reindex(columns=orden_meses, fill_value=0)

    # Restablecer el índice para que 'Componente' sea una columna
    df_resultado = df_resultado.reset_index()
    # Renombrar la columna 'index' a 'Componente'

    df_resultado = df_resultado.rename(columns={"index": "Componente"})
    df_resultado = df_resultado.rename_axis(None, axis=1)

    # Borro componentes no utilizados en modulo
    df_resultado = df_resultado.loc[~(df_resultado.drop(columns=['Componente']).sum(axis=1) == 0)]
    df_resultado = df_resultado.reset_index(drop=True)

    return df_resultado

def export_to_csv(df, name="output_data"):
    """
    Crea archivo csv en escritorio segun df recibido
    """
    output_file = os.path.join(url_desktop, f"{name}.csv")
    try:
        df.to_csv(output_file, index=False, encoding="utf-8-sig", sep=";")
        print(f"Archivo CSV creado: ./{name}.csv")
    except PermissionError:
        print(f"Error: No se pudo guardar el archivo '{name}.csv'. Verifica que no esté abierto en otro programa.")
    except Exception as e:
        print(f"Error inesperado al guardar el archivo: {e}")

def strip_columns(df):
    """
    Aplica .strip() a todas las celdas de texto en cada columna del DataFrame.
    """
    for col in df.columns:
        if df[col].dtype == 'object':  # Verifica si la columna contiene texto
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df


def menu(df):
    """
    Menú interactivo para ejecutar funciones con parámetros solicitados al usuario.
    """
    while True:
        # Mostrar el menú principal
        opcion = questionary.select(
            "Selecciona una opción:",
            choices=[
                "1. Filtrar por cantidad de registros (view_serie)",
                "2. Filtrar por número de serie (view_modulo)",
                "3. Ver componentes utilizados (por año o por mes)",
                "4. Salir"
            ]
        ).ask()

        if opcion == "1. Filtrar por cantidad de registros (view_serie)":
            # Solicitar el parámetro `cant_reg` al usuario
            cant_reg = questionary.text("Ingrese la cantidad mínima de registros:").ask()
            try:
                cant_reg = int(cant_reg)
                max_reg = df['Número de serie'].value_counts().max()
                view_serie(df, cant_reg=cant_reg)
            except ValueError:
                print(f"Por favor, ingrese un número válido. Máxima cantidad de registros: {max_reg}")

        elif opcion == "2. Filtrar por número de serie (view_modulo)":
            # Generar lista de números de serie únicos
            numeros_de_serie = sorted(df['Número de serie'].unique().tolist())

            # Solicitar al usuario que seleccione un número de serie
            n_serie = questionary.select(
                "Seleccione un número de serie:",
                choices=numeros_de_serie
            ).ask()

            # Llamar a la función view_modulo con el número de serie seleccionado
            view_modulo(df, n_serie=n_serie)

        elif opcion == "3. Ver componentes utilizados (por año o por mes)":
            # Submenú para elegir entre año o mes
            sub_opcion = questionary.select(
                "Seleccione una opción:",
                choices=[
                    "1. Filtrar por año",
                    "2. Filtrar por mes",
                    "3. Volver al menú principal"
                ]
            ).ask()

            if sub_opcion == "1. Filtrar por año":
                # Llamar al método view_component_used
                print("\nGenerando resumen de componentes utilizados por año...\n")
                view_component_used(df)

            elif sub_opcion == "2. Filtrar por mes":
                # Obtener lista de años disponibles
                list_anios = sorted(df['Año'].unique().tolist())

                # Solicitar al usuario que seleccione un año
                year = questionary.select(
                    "Seleccione un año:",
                    choices=[str(anio) for anio in list_anios]
                ).ask()

                # Imprimir un mensaje con el año seleccionado
                print(f"\nResumen de componentes utilizados por mes del año {year}.\n")
                df_month = components_used_by_month(df, "PWU", int(year))
                if df_month.empty:
                    print(f"Sin registros de componentes en el año {year}")
                else:
                    print(df_month)

            elif sub_opcion == "3. Volver al menú principal":
                continue

        elif opcion == "4. Salir":
            print("Saliendo del programa...")
            break

        else:
            print("Opción no válida. Por favor, selecciona una opción válida.")



if __name__ == "__main__":
    
    # Ingesta de datos y pre-procesamiento
    df = extract_xlsx()
    df = strip_columns(df)  
    # Transformacion datos nulos
    df = df.fillna("Sin registro")
    # Filtrado y extraccion de datos relevantes
    df = enrich_dataframe(df)
    # Elimina columnas innecesarias
    df = df.drop(list_components_pwu + list_components_bch, axis=1)

    print(df)
    print(df.info())

    #menu(df)

    """
    # Datos relevantes a cargar por usuario
    cant_reg_of_dfs = 3
    name_modulo = "DA30664"
    name_csv = f"Informe_de_filtrado_Mitsubishi"

    # Testeo de dfs
    view_serie(df, cant_reg= cant_reg_of_dfs)
    view_modulo(df, name_modulo)

    # Guardar el DataFrame en un archivo CSV
    export_to_csv(df, name=name_csv)   
    """


    
