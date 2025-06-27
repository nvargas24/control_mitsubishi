from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import calendar
import locale

import pandas as pd
from dotenv import load_dotenv
import os
import re

import questionary

# Configurar el idioma a español
locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252') 

load_dotenv('urls_sarmiento.env')
url_file_mit = os.getenv("URL_MITSUBISHI")
url_desktop = os.getenv("URL_DESKTOP")

list_components_bch = ['IGB1 1', 'IGB1 2', 'IGB1', 'IGB2', 'DB1', 'DB2']
list_components_pwu = ['IGD5 U', 'IGD5 V', 'IGD5 W', 'IGU', 'IGV', 'IGW', 'IGX', 'IGY', 'IGZ']

list_formaciones = ['RC01', 'RC02', 'RC03', 'RC04', 'RC05', 'RC06', 'RC07', 'RC08',
                    'RC09', 'RC10', 'RC11', 'RC12', 'RC13', 'RC14', 'RC15', 'RC16',
                    'RC17', 'RC18', 'RC19', 'RC20', 'RC21', 'RC22', 'RC23', 'RC24', 'RC25'
                    ]
list_coches = ['M1-1', 'M2-1', 'M1-2', 'M2-2', 'M3', 'M4']

def extract_xlsx():
    # Extraccion de URL del archivo xlsx desde un .env
    URL = os.path.join(url_file_mit, "FALLAS VVVF MITSUBISHI 20240129_V1.0.xlsx")

    # Carga de archivo xlsx
    wb = load_workbook(URL)
    ws = wb.active

    try:
        df_test = pd.read_excel(URL, sheet_name=ws.title, header=None)
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")

    start_table = "A1"
    end_table = "X1"

    start_row, start_col = coordinate_to_tuple(start_table)
    _, end_col = coordinate_to_tuple(end_table)

    # Ubico tabla en la hoja activa y obtengo un objeto generator
    rows = ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row, 
        min_col=start_col, 
        max_col=end_col,
    )

    # Itero sobre generator para cargar valores en una lista de listas
    data = []

    for row in rows:
        data.append([cell.value for cell in row])

    # Descrimino entre encabezado y datos
    headers = data[0]
    rows_data = data[1:]

    # Creo dataframe
    df = pd.DataFrame(rows_data, columns=headers)    

    # Identifico y filtro solo filas validas - en base a num de registro
    df["Num"] = pd.to_numeric(df["N"], errors="coerce")
    df_clean = df.dropna(subset=["Num"])
    df_clean = df_clean.drop("Num", axis=1)

    df_clean = df_clean.reset_index(drop=True)

    return df_clean

def filter_re_code(text):
    """
    Filtra textos que contengan la estructura 'RE****'.
    Si no se encuentra, devuelve 'Sin registro'.
    """
    match = re.search(r'\bRE\d{4}\b', text)
    if match:
        return match.group(0)
    else:
        return "Sin registro"

def enrich_dataframe(df):
    """
    Transforma e interpreta datos de tabla original para posterior analisis
    """
    # Extracción de num RE
    df['Cod_Rep'] = df['UBICACIÓN ACTUAL'].apply(filter_re_code)

    # Extracción de Tipo coche de una columna especifica
    serie_aux = df['Coche'].str.split(" ", expand=True)
    serie_aux.columns = ['Tipo coche', 'Num coche']
    serie_aux['Tipo coche'] = serie_aux['Tipo coche'].apply(lambda x: x if x in list_coches else "Sin registro") # Deteccion de campos no validos y reemplazo
    serie_aux = serie_aux.drop('Num coche', axis=1)    
    df = pd.concat([df, serie_aux], axis=1)
    df = df.drop('Coche', axis=1)

    # Ajuste de formato para 'Formación'
    df['Formación'] = df['Formación'].apply(lambda x: x if isinstance(x, (int, float)) else 0)
    df['Formación'] = df['Formación'].apply(lambda x: str(int(x)).zfill(2) if isinstance(x, (int, float)) else x)
    df['Formación'] = df['Formación'].apply(lambda x: f'RC{x}')
    df['Formación'] = df['Formación'].apply(lambda x: 'Sin registro' if x=="RC00" else x )

    # Interpretación de estados por referencia
    df = mapping_state(df) 

    return df

def export_to_csv(df, name="output_data"):
    """
    Crea archivo csv en escritorio segun df recibido
    """
    output_file = os.path.join(url_desktop, f"{name}.csv")
    try:
        df.to_csv(output_file, index=False, encoding="utf-8-sig", sep=";")
        print(f"Archivo CSV creado: ./{name}.csv")
    except PermissionError:
        print(f"Error: No se pudo guardar el archivo '{name}.csv'. Verifica que no esté abierto en otro programa.")
    except Exception as e:
        print(f"Error inesperado al guardar el archivo: {e}")

def strip_columns(df):
    """
    Elimina espacios en blanco tanto al inicio como al final de str
    """
    for col in df.columns:
        if df[col].dtype == 'object':  # Verifica si la columna contiene texto
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df

def format_df(df):
    mapping_format_df = {
        "N": "UInt16",
        "Fecha de falla": "datetime64[ns]",
        "Formación": "string",
        "Tipo coche": "string",
        "Unidad en falla": "string",
        "Número de serie": "string",
        "ESTADO ACTUAL": "string",
        "UBICACIÓN ACTUAL": "string",
        "GPS": "string",
        "Cod_Rep": "string"
    }

    df = df.astype(mapping_format_df)

    return df

def mapping_state(df):
    """
    Reemplaza estado a su versión completa
    """
    mapping = {
            "R": "Reparado",
            "D": "Desguase",
            "P": "Pediente",
            "SE": "Sin evaluación",
            "J": "Japón"
        }

    df['ESTADO ACTUAL'] = df['ESTADO ACTUAL'].replace(mapping).infer_objects(copy=False)
        
    return df


if __name__ == "__main__":
    
    # Ingesta de datos y pre-procesamiento
    df = extract_xlsx()
    df = strip_columns(df)  
    # Transformacion datos nulos
    df = df.fillna("Sin registro")
    # Filtrado y extraccion de datos relevantes - sin modificar
    df = enrich_dataframe(df) 
    df_filter = format_df(df)
    # Elimina columnas innecesarias -- DEPENDE DF
    df_filter = df_filter.drop(list_components_pwu + list_components_bch, axis=1)

    print(df_filter)
    print(df_filter.info())  

    # Exporta en formato csv
    export_to_csv(df_filter, "registros_fallas_mitsubishi")
