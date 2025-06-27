from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple

import calendar
import locale

import pandas as pd
from dotenv import load_dotenv
import os
import re

import questionary

# Configurar el idioma a español
locale.setlocale(locale.LC_TIME, 'Spanish_Spain.1252') 

load_dotenv('urls_sarmiento.env')
url_file_mit = os.getenv("URL_MITSUBISHI")
url_desktop = os.getenv("URL_DESKTOP")

list_components_bch = ['IGB1 1', 'IGB1 2', 'IGB1', 'IGB2', 'DB1', 'DB2']
list_components_pwu = ['IGD5 U', 'IGD5 V', 'IGD5 W', 'IGU', 'IGV', 'IGW', 'IGX', 'IGY', 'IGZ']

list_formaciones = ['RC01', 'RC02', 'RC03', 'RC04', 'RC05', 'RC06', 'RC07', 'RC08',
                    'RC09', 'RC10', 'RC11', 'RC12', 'RC13', 'RC14', 'RC15', 'RC16',
                    'RC17', 'RC18', 'RC19', 'RC20', 'RC21', 'RC22', 'RC23', 'RC24', 'RC25'
                    ]
list_coches = ['M1-1', 'M2-1', 'M1-2', 'M2-2', 'M3', 'M4']

def extract_xlsx():
    # Extraccion de URL del archivo xlsx desde un .env
    URL = os.path.join(url_file_mit, "FALLAS VVVF MITSUBISHI 20240129_V1.0.xlsx")

    # Carga de archivo xlsx
    wb = load_workbook(URL)
    ws = wb.active

    try:
        df_test = pd.read_excel(URL, sheet_name=ws.title, header=None)
    except Exception as e:
        print(f"Error al cargar el archivo: {e}")

    start_table = "A1"
    end_table = "X1"

    start_row, start_col = coordinate_to_tuple(start_table)
    _, end_col = coordinate_to_tuple(end_table)

    # Ubico tabla en la hoja activa y obtengo un objeto generator
    rows = ws.iter_rows(
        min_row=start_row,
        max_row=ws.max_row, 
        min_col=start_col, 
        max_col=end_col,
    )

    # Itero sobre generator para cargar valores en una lista de listas
    data = []

    for row in rows:
        data.append([cell.value for cell in row])

    # Descrimino entre encabezado y datos
    headers = data[0]
    rows_data = data[1:]

    # Creo dataframe
    df = pd.DataFrame(rows_data, columns=headers)    

    # Identifico y filtro solo filas validas - en base a num de registro
    df["Num"] = pd.to_numeric(df["N"], errors="coerce")
    df_clean = df.dropna(subset=["Num"])
    df_clean = df_clean.drop("Num", axis=1)

    df_clean = df_clean.reset_index(drop=True)

    return df_clean

def calculate_time_between_failures(df):
    # Calcular la diferencia en días entre fechas consecutivas
    df['Tiempo entre fallas (días)'] = df['Fecha de falla'].diff().dt.days
    # Reemplazar NaN (primer registro) con 0
    df['Tiempo entre fallas (días)'] = df['Tiempo entre fallas (días)'].fillna(0).astype(int)

    return df

def search_serie(df, num_serie):
    df_filter = df.copy()

    try:
        df_filter = df_filter[df_filter["Número de serie"] == num_serie]

        # Identifico modulo segun numero de serie solicitado
        if not df_filter.empty:
            modulo = df_filter.iloc[0]['Unidad en falla']
        else:
            raise ValueError(f"No se encontró el número de serie: {num_serie}")

        # Borra columnas no relevantes
        df_filter = df_filter.drop(['Número de serie', 'Unidad en falla','UBICACIÓN ACTUAL', 'GPS', 'componentes_reemplazados'], axis=1)

        # Calculo de lapso de tiempo entre ingresos
        df_filter = calculate_time_between_failures(df_filter)

        return df_filter, modulo

    except ValueError as ve:
        print(ve)
        return pd.DataFrame(), None  # Retorna un DataFrame vacío y `None` como módulo
    except KeyError as ke:
        print(f"Error: Falta una columna esperada en el DataFrame: {ke}")
        return pd.DataFrame(), None
    except Exception as e:
        print(f"Error inesperado: {e}")
        return pd.DataFrame(), None

def filter_by_type(df, type):
    df_filter = df.copy()

    df_filter = df_filter[df_filter["Unidad en falla"] == type]
    #df_filter = df_filter.drop("Unidad en falla", axis=1)

    if type=='BCH':
        df_filter = df_filter.drop(list_components_pwu, axis=1)
    elif type=='PWU':
        df_filter = df_filter.drop(list_components_bch, axis=1)

    return df_filter

def filter_re_code(text):
    """
    Filtra textos que contengan la estructura 'RE****'.
    Si no se encuentra, devuelve 'Sin registro'.
    """
    match = re.search(r'\bRE\d{4}\b', text)
    if match:
        return match.group(0)
    else:
        return "Sin registro"

def enrich_dataframe(df):
    df['Cod_Rep'] = df['UBICACIÓN ACTUAL'].apply(filter_re_code)
    # Obtiene Tipo coche de una columna especifica
    serie_aux = df['Coche'].str.split(" ", expand=True)
    serie_aux.columns = ['Tipo coche', 'Num coche']
    # Deteccion de campos no validos y reemplazo
    serie_aux['Tipo coche'] = serie_aux['Tipo coche'].apply(lambda x: x if x in list_coches else "Sin registro")
    serie_aux = serie_aux.drop('Num coche', axis=1)
    #print(serie_aux)
    df = pd.concat([df, serie_aux], axis=1)
    df = df.drop('Coche', axis=1)

    # Ajuste de formato para 'Formación'
    df['Formación'] = df['Formación'].apply(lambda x: x if isinstance(x, (int, float)) else 0)
    df['Formación'] = df['Formación'].apply(lambda x: str(int(x)).zfill(2) if isinstance(x, (int, float)) else x)
    df['Formación'] = df['Formación'].apply(lambda x: f'RC{x}')
    df['Formación'] = df['Formación'].apply(lambda x: 'Sin registro' if x=="RC00" else x )

    return df

def view_serie(df, cant_reg=0):
    df_serie = df.copy()

    list_cant_sup_reg = []
    num_reg_max = 0

    for n_serie in df_serie['Número de serie'].unique():
        try:
            df_filtered, modulo = search_serie(df_serie, f"{n_serie}")
            # ##### FILTRAR SI ES PWU O BCH  ***** en todas las lecturas

            df_filtered = df_filtered.drop(['Año', 'Mes'], axis=1)
            # Procesar solo si no hay errores en search_serie
            num_reg = df_filtered.shape[0]
            
            if num_reg > num_reg_max:
                num_reg_max = num_reg

            if num_reg == cant_reg:
                print(f"\n ********************** Equipo {modulo} - {n_serie} ****************************")
                print(df_filtered)

            if num_reg > cant_reg:
                list_cant_sup_reg.append(n_serie)

        except ValueError as ve:
            print(ve)
        except Exception as e:
            print(f"Error inesperado: {e}")

    if cant_reg > num_reg_max:
        raise ValueError(f"El valor de cant_reg ({cant_reg}) es mayor que el número máximo de registros ({num_reg}) para el número de serie: {n_serie}")

    if list_cant_sup_reg:
        str_sup_reg = ", ".join(list_cant_sup_reg)

        print(f" *****  ADV: Hay modulos con un numero mayor de {cant_reg} registro/s:")
        print(f" *****  {str_sup_reg}")

def view_modulo(df, n_serie):
    df_serie = df.copy()

    df_filtered, modulo = search_serie(df_serie, f"{n_serie}")
    if not df_filtered.empty and modulo is not None:
        print(f"\n ********************** Equipo {modulo} - {n_serie} ****************************")
        print(df_filtered)

def view_component_used(df):
    df_comp = df.copy()
    
    df_resume = None  # Inicializar el DataFrame resumen
    #list_modulos = df_comp["Unidad en falla"].unique()
    list_anios = df_comp['Año'].unique()
    list_modulo = df_comp['Unidad en falla'].unique()

    for modulo in list_modulo:
        for year in list_anios:
            try:
                # Generar el DataFrame para la combinación actual
                df_result = components_used_by_month(df_comp, modulo, year)
                # Agregar la columna 'Total' con la suma de todas las columnas de meses
                df_result[year] = df_result.drop(columns=['Componente']).sum(axis=1)
                # Mantener solo las columnas 'Componente' y el total del año
                df_result = df_result[['Componente', year]]
                
                # Combinar con el DataFrame resumen
                if df_resume is None:
                    df_resume = df_result  # Inicializar con el primer DataFrame
                else:
                    # Combinar basándose en la columna 'Componente'
                    df_resume = pd.merge(df_resume, df_result, on='Componente', how='outer')

            except Exception as e:
                print(f"Error al procesar módulo '{modulo}' y año '{year}': {e}")

        df_resume = df_resume.apply(
            lambda col: col.fillna(0).astype(int) if col.dtype in ['float64', 'int64'] else col
        )
        print(f"\n ********************** Equipos {modulo} ****************************")
        print(df_resume)  
        df_resume = None

def view_resume_formacion(df_original):
    df_comp = df_original.copy()
    list_modulo = df_comp['Unidad en falla'].unique()

    for modulo in list_modulo:
        df_resume = resume_formacion(df_comp, modulo)
        print(f"\n ********************** Equipos {modulo} ****************************")
        print(df_resume)  
        df_resume = None

def create_empty_formacion_df():
    """
    Crea un DataFrame con la columna 'Formación' y columnas adicionales para cada elemento de 'list_coches'.
    Las filas de 'Formación' se llenan con los valores de 'list_formaciones', y las demás columnas se inicializan con None.
    """
    # Crear un diccionario con 'Formación' y columnas de 'list_coches'
    data = {'Formación': list_formaciones}
    for coche in list_coches:
        data[coche] = [r"-"] * len(list_formaciones)  # Inicializar con None

    # Crear el DataFrame
    df = pd.DataFrame(data)
    return df

def resume_formacion(df_original, modulo):
    df = df_original.copy()
    df_formaciones = create_empty_formacion_df()
    dic_historial_serie = {}

    df = df[['Fecha de falla', 'Número de serie', 'Formación', 'Tipo coche', 'Unidad en falla']]
    df = df[df['Unidad en falla']== modulo]
    df = df.drop('Unidad en falla', axis =1)
    df = df.sort_values(by='Fecha de falla', ascending=True).sort_index(ascending=False).reset_index(drop=True)

    # Historial de equipos en formacion --- obs: antes del filtrar ordenar por fecha
    for formacion in list_formaciones:
        aux = df.loc[df['Formación']== formacion, 'Número de serie']

        str_aux = aux.str.cat(sep=", ")
        dic_historial_serie[formacion] = str_aux

    for form in list_formaciones:
        historial_mod = dic_historial_serie[form]
        if historial_mod:
            df_formaciones.loc[df_formaciones['Formación']==form, 'Historial'] = historial_mod
        else:
            df_formaciones.loc[df_formaciones['Formación']==form, 'Historial'] = "-"

    # Registro de modulo actual segun coche y formacion
    for form in list_formaciones:
        for coche in list_coches:
            df_filtrado = df[(df['Formación'] == form) & (df['Tipo coche'] == coche)]
            if not df_filtrado.empty:
                value = df_filtrado.iloc[0]['Número de serie']
            else:
                value = "-"  
            
            df_formaciones.loc[df_formaciones['Formación']==form, coche] = value

    # Registro de ultima falla
    df_last_reg = df.groupby(['Formación'])['Fecha de falla'].max().reset_index()
    df_formaciones = df_formaciones.merge(df_last_reg, on='Formación', how='left')
    df_formaciones.rename(columns={'Fecha de falla': 'Fecha ultima falla'}, inplace=True)
    
    return df_formaciones


def export_to_csv(df, name="output_data"):
    """
    Crea archivo csv en escritorio segun df recibido
    """
    output_file = os.path.join(url_desktop, f"{name}.csv")
    try:
        df.to_csv(output_file, index=False, encoding="utf-8-sig", sep=";")
        print(f"Archivo CSV creado: ./{name}.csv")
    except PermissionError:
        print(f"Error: No se pudo guardar el archivo '{name}.csv'. Verifica que no esté abierto en otro programa.")
    except Exception as e:
        print(f"Error inesperado al guardar el archivo: {e}")

def strip_columns(df):
    """
    Elimina espacios en blanco tanto al inicio como al final de str
    """
    for col in df.columns:
        if df[col].dtype == 'object':  # Verifica si la columna contiene texto
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    return df

def format_df(df):
    mapping_format_df = {
        "N": "UInt16",
        "Fecha de falla": "datetime64[ns]",
        "Formación": "string",
        "Tipo coche": "string",
        "Unidad en falla": "string",
        "Número de serie": "string",
        "ESTADO ACTUAL": "string",
        "UBICACIÓN ACTUAL": "string",
        "GPS": "string",
        "Cod_Rep": "string"
    }

    df = df.astype(mapping_format_df)

    return df


def mapping_state(df):
    """
    Reemplaza estado a su versión completa
    """
    mapping = {
            "R": "Reparado",
            "D": "Desguase",
            "P": "Pediente",
            "SE": "Sin evaluación",
            "J": "Japón"
        }

    df['ESTADO ACTUAL'] = df['ESTADO ACTUAL'].replace(mapping).infer_objects(copy=False)
        
    return df


if __name__ == "__main__":
    
    # Ingesta de datos y pre-procesamiento
    df = extract_xlsx()
    df = strip_columns(df)  
    # Transformacion datos nulos
    df = df.fillna("Sin registro")
    # Filtrado y extraccion de datos relevantes - sin modificar
    df = enrich_dataframe(df) 
    df_filter = format_df(df)
    # Elimina columnas innecesarias -- DEPENDE DF
    df_filter = df_filter.drop(list_components_pwu + list_components_bch, axis=1)

    print(df_filter)
    print(df_filter.info())  

